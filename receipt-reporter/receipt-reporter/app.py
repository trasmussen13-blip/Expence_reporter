from __future__ import annotations

import base64
import functools
import hmac
import io
import json
import os
import re
import secrets
import shutil
import sqlite3
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
import requests
from werkzeug.utils import secure_filename


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "receipt_data"
RECEIPT_DIR = DATA_DIR / "receipts"
EXPORT_DIR = DATA_DIR / "exports"
USER_TEMPLATE_DIR = DATA_DIR / "user_templates"
DB_PATH = DATA_DIR / "receipts.sqlite3"
TEMPLATE_PATH = DATA_DIR / "base_template.xlsx"
SEED_TEMPLATE = (
    ROOT
    / ".local"
    / "conversation-workspace"
    / "files"
    / "attached_assets"
    / "98d8be45-7805-4237-a46a-34c9a662e094_1788530326600.xlsx"
)

ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg"}
SHEET_NAME = "1"
DATA_START_ROW = 4
CATEGORY_COLUMNS = {
    "vehicle": "C",
    "travel": "D",
    "hotel": "E",
    "other_operating_expense": "F",
    "hospitality": "G",
    "equipment": "F",
}
CATEGORY_LABELS = {
    "vehicle": "#680111/#6430 · Køretøj",
    "travel": "#620951/#6631 · Taxi / fly / tog",
    "hotel": "#620951/#6630 · Hotel",
    "other_operating_expense": "#655761/#6940 · Øvrig driftsudgift",
    "hospitality": "#620951/#6613 · Forplejning",
    "equipment": "#655761/#6940 · Udstyr",
}
DEFAULT_CATEGORY = "other_operating_expense"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.secret_key = os.getenv("SESSION_SECRET") or secrets.token_hex(32)
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.getenv("COOKIE_SECURE", "").lower() in {"1", "true", "yes"}

OAUTH_PROVIDERS = {
    "google": {
        "label": "Google",
        "client_id_env": "GOOGLE_CLIENT_ID",
        "client_secret_env": "GOOGLE_CLIENT_SECRET",
        "authorize_url": "https://accounts.google.com/o/oauth2/v2/auth",
        "token_url": "https://oauth2.googleapis.com/token",
        "scope": "openid email profile",
    },
    "microsoft": {
        "label": "Microsoft",
        "client_id_env": "MICROSOFT_CLIENT_ID",
        "client_secret_env": "MICROSOFT_CLIENT_SECRET",
        "authorize_url": "https://login.microsoftonline.com/common/oauth2/v2.0/authorize",
        "token_url": "https://login.microsoftonline.com/common/oauth2/v2.0/token",
        "scope": "openid profile email User.Read",
    },
}


def ensure_storage() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    RECEIPT_DIR.mkdir(exist_ok=True)
    EXPORT_DIR.mkdir(exist_ok=True)
    USER_TEMPLATE_DIR.mkdir(exist_ok=True)
    if not TEMPLATE_PATH.exists() and SEED_TEMPLATE.exists():
        shutil.copy2(SEED_TEMPLATE, TEMPLATE_PATH)
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                provider TEXT NOT NULL,
                provider_subject TEXT NOT NULL,
                email TEXT NOT NULL DEFAULT '',
                display_name TEXT NOT NULL DEFAULT '',
                report_master_name TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(provider, provider_subject)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS receipts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                stored_filename TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                receipt_date TEXT,
                vendor TEXT NOT NULL DEFAULT '',
                category TEXT NOT NULL DEFAULT 'other_operating_expense',
                amount REAL,
                vat REAL,
                currency TEXT NOT NULL DEFAULT 'DKK',
                extraction_note TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
        receipt_columns = {
            row[1] for row in conn.execute("PRAGMA table_info(receipts)").fetchall()
        }
        if "user_id" not in receipt_columns:
            conn.execute("ALTER TABLE receipts ADD COLUMN user_id INTEGER")
        conn.commit()


def user_by_id(user_id: int) -> dict[str, Any] | None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return dict(row) if row else None


def current_user() -> dict[str, Any] | None:
    user_id = session.get("user_id")
    if not user_id:
        return None
    try:
        return user_by_id(int(user_id))
    except (TypeError, ValueError):
        return None


def login_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if current_user() is None:
            return redirect(url_for("index"))
        return view(*args, **kwargs)

    return wrapped


def db_rows(user_id: int) -> list[dict[str, Any]]:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT * FROM receipts
            WHERE user_id = ?
            ORDER BY COALESCE(receipt_date, ''), id
            """,
            (user_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def db_one(user_id: int, receipt_id: int) -> dict[str, Any] | None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT * FROM receipts WHERE id = ? AND user_id = ?",
            (receipt_id, user_id),
        ).fetchone()
    return dict(row) if row else None


def upsert_user(provider: str, profile: dict[str, Any]) -> dict[str, Any]:
    subject = str(profile.get("subject") or "").strip()
    if not subject:
        raise ValueError("Loginudbyderen returnerede ikke et bruger-id.")
    email = str(profile.get("email") or "").strip()
    display_name = str(profile.get("display_name") or email or "Bruger").strip()
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO users (provider, provider_subject, email, display_name, report_master_name)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(provider, provider_subject) DO UPDATE SET
                email = excluded.email,
                display_name = excluded.display_name,
                updated_at = CURRENT_TIMESTAMP
            """,
            (provider, subject, email, display_name, display_name),
        )
        user_id = conn.execute(
            "SELECT id FROM users WHERE provider = ? AND provider_subject = ?",
            (provider, subject),
        ).fetchone()[0]
        conn.commit()
    return user_by_id(user_id)


def user_template_path(user_id: int) -> Path:
    return USER_TEMPLATE_DIR / f"user_{user_id}.xlsx"


def oauth_config(provider: str) -> dict[str, Any]:
    config = OAUTH_PROVIDERS.get(provider)
    if not config:
        raise ValueError("Ukendt loginudbyder.")
    client_id = os.getenv(config["client_id_env"], "").strip()
    client_secret = os.getenv(config["client_secret_env"], "").strip()
    if not client_id or not client_secret:
        raise RuntimeError(
            f"{config['label']}-login er ikke konfigureret på denne server."
        )
    return {**config, "client_id": client_id, "client_secret": client_secret}


def oauth_redirect_uri(provider: str) -> str:
    base_url = os.getenv("OAUTH_REDIRECT_BASE_URL", "").strip().rstrip("/")
    if not base_url:
        base_url = request.url_root.rstrip("/")
    return f"{base_url}/auth/callback/{provider}"


def oauth_profile(provider: str, access_token: str) -> dict[str, str]:
    headers = {"Authorization": f"Bearer {access_token}"}
    if provider == "google":
        response = requests.get(
            "https://openidconnect.googleapis.com/v1/userinfo",
            headers=headers,
            timeout=20,
        )
        response.raise_for_status()
        profile = response.json()
        return {
            "subject": str(profile.get("sub") or ""),
            "email": str(profile.get("email") or ""),
            "display_name": str(profile.get("name") or profile.get("email") or ""),
        }

    response = requests.get(
        "https://graph.microsoft.com/v1.0/me",
        headers=headers,
        timeout=20,
    )
    response.raise_for_status()
    profile = response.json()
    return {
        "subject": str(profile.get("id") or ""),
        "email": str(profile.get("mail") or profile.get("userPrincipalName") or ""),
        "display_name": str(profile.get("displayName") or profile.get("mail") or ""),
    }


def render_login(error: str = ""):
    configured = {}
    for provider, config in OAUTH_PROVIDERS.items():
        configured[provider] = bool(
            os.getenv(config["client_id_env"], "").strip()
            and os.getenv(config["client_secret_env"], "").strip()
        )
    return render_template(
        "login.html",
        providers=OAUTH_PROVIDERS,
        configured=configured,
        error=error,
    )


def parse_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (float, int)):
        return round(float(value), 2)
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def normalize_date(value: Any) -> str:
    if not value:
        return ""
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    match = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", text)
    if match:
        day, month, year = match.groups()
        year = f"20{year}" if len(year) == 2 else year
        try:
            return date(int(year), int(month), int(day)).isoformat()
        except ValueError:
            return ""
    return ""


def normalize_category(value: Any) -> str:
    key = str(value or "").strip().lower()
    aliases = {
        "vehicle": "vehicle",
        "car": "vehicle",
        "transport": "travel",
        "travel": "travel",
        "taxi": "travel",
        "flight": "travel",
        "train": "travel",
        "hotel": "hotel",
        "accommodation": "hotel",
        "hospitality": "hospitality",
        "food": "hospitality",
        "restaurant": "hospitality",
        "equipment": "equipment",
        "other": "other_operating_expense",
        "other_operating_expense": "other_operating_expense",
    }
    return aliases.get(key, DEFAULT_CATEGORY)


def normalized_vat(amount: Any, vat: Any) -> float | None:
    amount_value = parse_number(amount)
    vat_value = parse_number(vat)
    if vat_value is None:
        return None
    vat_value = max(vat_value, 0)
    if amount_value is not None:
        vat_value = min(vat_value, amount_value)
    return round(vat_value, 2)


def image_for_ai(image_path: Path) -> tuple[str, str]:
    """Return the original image as a base64 Ollama Vision payload."""
    suffix = image_path.suffix.lower().lstrip(".")
    mime = "image/jpeg" if suffix in {"jpg", "jpeg"} else f"image/{suffix}"
    return mime, base64.b64encode(image_path.read_bytes()).decode("ascii")


def clean_json_response(raw: str) -> str:
    """Strip common Markdown wrappers and isolate the JSON object."""
    cleaned = (raw or "").strip()
    cleaned = re.sub(r"^\s*```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s*```\s*$", "", cleaned)
    match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
    return match.group(0) if match else cleaned


def extract_receipt(image_path: Path) -> dict[str, Any]:
    _, encoded = image_for_ai(image_path)
    prompt = """
Læs denne kvittering. Returnér KUN ét råt, gyldigt JSON-objekt uden
Markdown, kodeblokke, forklaring eller ekstra tekst.
{
  "date": "YYYY-MM-DD eller tom streng",
  "merchant": "butik/leverandør",
  "total": 0.0,
  "vat": 0.0
}
Brug det samlede beløb inkl. moms som total. Brug 0.0, hvis moms ikke kan
aflæses. Brug altid præcis disse fire nøgler. Datoen skal være YYYY-MM-DD
eller en tom streng. Gæt ikke på oplysninger, der ikke kan aflæses.
""".strip()

    try:
        ollama_host = os.getenv("OLLAMA_HOST", "http://localhost:11434").rstrip("/")
        response = requests.post(
            f"{ollama_host}/api/chat",
            json={
                "model": "llama3.2-vision",
                "messages": [
                    {
                        "role": "user",
                        "content": prompt,
                        "images": [encoded],
                    }
                ],
                "stream": False,
                "format": "json",
                "options": {"temperature": 0},
            },
            timeout=120,
        )
        response.raise_for_status()
        raw = response.json().get("message", {}).get("content", "")
        data = json.loads(clean_json_response(raw))
        return {
            "date": normalize_date(data.get("date")),
            "vendor": str(data.get("merchant") or "").strip(),
            "category": DEFAULT_CATEGORY,
            "amount": parse_number(data.get("total")),
            "vat": parse_number(data.get("vat")),
            "note": "",
        }
    except Exception as exc:
        return {
            "date": "",
            "vendor": "",
            "category": DEFAULT_CATEGORY,
            "amount": None,
            "vat": None,
            "note": f"Lokal AI-analyse fejlede: {str(exc)[:180]}",
        }


def find_total_row(ws) -> int:
    for row in range(DATA_START_ROW, ws.max_row + 1):
        if str(ws.cell(row, 2).value or "").strip().lower() == "total":
            return row
    return ws.max_row + 1


def copy_row_format(ws, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = source._style.copy()
        if source.number_format:
            target.number_format = source.number_format
        target.font = Font(
            name=source.font.name,
            sz=source.font.sz,
            b=source.font.b,
            i=source.font.i,
            color=source.font.color,
            vertAlign=source.font.vertAlign,
            underline=source.font.underline,
            strike=source.font.strike,
        )
        target.fill = PatternFill(
            fill_type=source.fill.fill_type,
            fgColor=source.fill.fgColor,
            bgColor=source.fill.bgColor,
        )
        target.border = Border(
            left=source.border.left,
            right=source.border.right,
            top=source.border.top,
            bottom=source.border.bottom,
            diagonal=source.border.diagonal,
            diagonal_direction=source.border.diagonal_direction,
            diagonalUp=source.border.diagonalUp,
            diagonalDown=source.border.diagonalDown,
            outline=source.border.outline,
            vertical=source.border.vertical,
            horizontal=source.border.horizontal,
        )
        target.alignment = Alignment(
            horizontal=source.alignment.horizontal,
            vertical=source.alignment.vertical,
            text_rotation=source.alignment.text_rotation,
            wrap_text=source.alignment.wrap_text,
            shrink_to_fit=source.alignment.shrink_to_fit,
            indent=source.alignment.indent,
        )
        target.protection = Protection(
            locked=source.protection.locked,
            hidden=source.protection.hidden,
        )


def populate_workbook(user_id: int, report_master_name: str) -> tuple[io.BytesIO, str]:
    template_path = user_template_path(user_id)
    if not template_path.exists():
        template_path = TEMPLATE_PATH
    if not template_path.exists():
        raise FileNotFoundError("Upload først den officielle Excel-skabelon.")
    if not report_master_name.strip():
        raise ValueError("Indtast rapportansvarligens navn før eksport.")

    workbook = load_workbook(template_path)
    ws = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    ws["A1"] = report_master_name.strip()
    total_row = find_total_row(ws)
    receipts = db_rows(user_id)
    available = [
        row
        for row in range(DATA_START_ROW, total_row)
        if all(ws.cell(row, col).value is None for col in range(1, 11))
    ]

    if len(receipts) > len(available):
        extra = len(receipts) - len(available)
        source_row = max(DATA_START_ROW, total_row - 1)
        ws.insert_rows(total_row, extra)
        for row in range(total_row, total_row + extra):
            copy_row_format(ws, source_row, row)
            ws.cell(row, 11).value = f"=SUM(C{row}:J{row})"
        total_row += extra
        available.extend(range(total_row - extra, total_row))

    target_rows = available[: len(receipts)]
    for row, receipt in zip(target_rows, receipts):
        for col in range(1, 12):
            if col != 11:
                ws.cell(row, col).value = None
        parsed_date = normalize_date(receipt.get("receipt_date"))
        if parsed_date:
            ws.cell(row, 1).value = datetime.strptime(parsed_date, "%Y-%m-%d").date()
        ws.cell(row, 2).value = receipt.get("vendor") or ""
        amount = receipt.get("amount")
        vat = normalized_vat(amount, receipt.get("vat")) or 0
        category = normalize_category(receipt.get("category"))
        if amount is not None:
            net = max(float(amount) - float(vat), 0)
            category_column = CATEGORY_COLUMNS.get(category, "F")
            ws[f"{category_column}{row}"].value = round(net, 2)
            if vat:
                ws.cell(row, 8).value = round(float(vat), 2)
        ws.cell(row, 11).value = f"=SUM(C{row}:J{row})"
        ws.cell(row, 1).number_format = "dd-mm-yyyy"

    # Keep the original total formulas, but extend their ranges when rows were inserted.
    for col in range(3, 12):
        letter = ws.cell(total_row, col).column_letter
        ws.cell(total_row, col).value = f"=SUM({letter}{DATA_START_ROW}:{letter}{total_row - 1})"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, f"kvitteringsrapport_{date.today().strftime('%Y-%m')}.xlsx"


def build_zip(user_id: int, report_master_name: str) -> tuple[io.BytesIO, str]:
    excel_bytes, excel_name = populate_workbook(user_id, report_master_name)
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(excel_name, excel_bytes.getvalue())
        for receipt in db_rows(user_id):
            path = RECEIPT_DIR / receipt["stored_filename"]
            if not path.exists():
                continue
            receipt_date = receipt.get("receipt_date") or "uden-dato"
            vendor = re.sub(r"[^A-Za-z0-9ÆØÅæøå-]+", "-", receipt.get("vendor") or "ukendt")
            amount = receipt.get("amount")
            amount_label = f"{amount:.2f}" if isinstance(amount, (int, float)) else "ukendt"
            extension = path.suffix.lower()
            name = f"{receipt_date}_{vendor}_{amount_label}{extension}"
            archive.write(path, f"Kvitteringer/{name}")
    zip_buffer.seek(0)
    return zip_buffer, f"kvitteringer_{date.today().strftime('%Y-%m')}.zip"


@app.get("/")
def index():
    ensure_storage()
    user = current_user()
    if user is None:
        return render_login()
    receipts = db_rows(user["id"])
    total = sum(float(row["amount"] or 0) for row in receipts)
    vat_total = sum(float(row["vat"] or 0) for row in receipts)
    return render_template(
        "index.html",
        receipts=receipts,
        user=user,
        total=total,
        vat_total=vat_total,
        template_exists=user_template_path(user["id"]).exists() or TEMPLATE_PATH.exists(),
        categories=CATEGORY_LABELS,
        category_columns=CATEGORY_COLUMNS,
        today=date.today().isoformat(),
    )


@app.get("/favicon.ico")
def favicon():
    return send_file(ROOT / "static" / "favicon.svg", mimetype="image/svg+xml")


@app.get("/auth/<provider>")
def start_login(provider: str):
    ensure_storage()
    try:
        config = oauth_config(provider)
    except (RuntimeError, ValueError) as exc:
        return render_login(str(exc)), 400
    state = secrets.token_urlsafe(32)
    session["oauth_state"] = state
    session["oauth_provider"] = provider
    params = {
        "client_id": config["client_id"],
        "redirect_uri": oauth_redirect_uri(provider),
        "response_type": "code",
        "scope": config["scope"],
        "state": state,
    }
    if provider == "google":
        params["access_type"] = "online"
        params["prompt"] = "select_account"
    else:
        params["response_mode"] = "query"
    return redirect(f"{config['authorize_url']}?{urlencode(params)}")


@app.get("/auth/callback/<provider>")
def finish_login(provider: str):
    expected_state = session.pop("oauth_state", "")
    expected_provider = session.pop("oauth_provider", "")
    state = request.args.get("state", "")
    if (
        not expected_state
        or not expected_provider
        or expected_provider != provider
        or not hmac.compare_digest(expected_state, state)
    ):
        return render_login("Login-sessionen udløb. Prøv igen."), 400
    code = request.args.get("code", "")
    if not code:
        return render_login("Login blev annulleret eller returnerede ingen kode."), 400
    try:
        config = oauth_config(provider)
        token_response = requests.post(
            config["token_url"],
            data={
                "code": code,
                "client_id": config["client_id"],
                "client_secret": config["client_secret"],
                "redirect_uri": oauth_redirect_uri(provider),
                "grant_type": "authorization_code",
            },
            timeout=20,
        )
        token_response.raise_for_status()
        access_token = token_response.json().get("access_token", "")
        if not access_token:
            raise RuntimeError("Loginudbyderen returnerede ikke en adgangstoken.")
        profile = oauth_profile(provider, access_token)
        user = upsert_user(provider, profile)
        session.clear()
        session["user_id"] = user["id"]
        return redirect(url_for("index"))
    except Exception as exc:
        return render_login(f"Login fejlede: {str(exc)[:180]}"), 400


@app.get("/logout")
def logout():
    session.clear()
    return redirect(url_for("index"))


@app.post("/profile")
@login_required
def update_profile():
    name = str(request.form.get("report_master_name") or "").strip()
    if not name:
        return redirect(url_for("index", profile_error="Indtast et navn til rapportmasteren."))
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            "UPDATE users SET report_master_name = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            (name[:120], int(session["user_id"])),
        )
        conn.commit()
    return redirect(url_for("index", profile_saved="1"))


@app.errorhandler(413)
def upload_too_large(_error):
    return redirect(
        url_for(
            "index",
            upload_error="Billedet er for stort. Vælg et billede under 50 MB.",
        )
    )


@app.post("/template")
@login_required
def upload_template():
    ensure_storage()
    user = current_user()
    uploaded = request.files.get("template")
    if not uploaded or not uploaded.filename:
        return redirect(url_for("index"))
    name = secure_filename(uploaded.filename)
    if not name.lower().endswith((".xlsx", ".xlsm")):
        return redirect(url_for("index"))
    uploaded.save(user_template_path(user["id"]))
    return redirect(url_for("index", template_saved="1"))


@app.post("/process")
@login_required
def process_receipts():
    ensure_storage()
    user = current_user()
    uploads = request.files.getlist("receipts")
    accepted = 0
    for uploaded in uploads:
        original = secure_filename(uploaded.filename or "")
        suffix = Path(original).suffix.lower().lstrip(".")
        if not original or suffix not in ALLOWED_IMAGE_EXTENSIONS:
            continue
        stored = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{accepted}.{suffix}"
        destination = RECEIPT_DIR / stored
        uploaded.save(destination)
        extracted = extract_receipt(destination)
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                """
                INSERT INTO receipts
                    (user_id, stored_filename, original_filename, receipt_date, vendor, category,
                     amount, vat, currency, extraction_note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'DKK', ?)
                """,
                (
                    user["id"],
                    stored,
                    original,
                    extracted["date"],
                    extracted["vendor"],
                    extracted["category"],
                    extracted["amount"],
                    extracted["vat"],
                    extracted["note"],
                ),
            )
            conn.commit()
        accepted += 1
    return redirect(url_for("index", processed=str(accepted)))


@app.patch("/receipt-actions/receipts/<int:receipt_id>")
@login_required
def update_receipt(receipt_id: int):
    ensure_storage()
    user = current_user()
    payload = request.get_json(silent=True) or {}
    category = normalize_category(payload.get("category"))
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            UPDATE receipts
            SET receipt_date = ?, vendor = ?, category = ?, amount = ?, vat = ?
            WHERE id = ? AND user_id = ?
            """,
            (
                normalize_date(payload.get("date")),
                str(payload.get("vendor") or "").strip(),
                category,
                parse_number(payload.get("amount")),
                normalized_vat(payload.get("amount"), payload.get("vat")),
                receipt_id,
                user["id"],
            ),
        )
        conn.commit()
    return jsonify({"ok": True})


@app.delete("/receipt-actions/receipts/<int:receipt_id>")
@login_required
def delete_receipt(receipt_id: int):
    ensure_storage()
    user = current_user()
    row = db_one(user["id"], receipt_id)
    if row:
        with sqlite3.connect(DB_PATH) as conn:
            conn.execute(
                "DELETE FROM receipts WHERE id = ? AND user_id = ?",
                (receipt_id, user["id"]),
            )
            conn.commit()
        path = RECEIPT_DIR / row["stored_filename"]
        if path.exists():
            path.unlink()
    return jsonify({"ok": True})


@app.get("/receipt-actions/export")
@login_required
def export_excel():
    ensure_storage()
    user = current_user()
    try:
        output, filename = populate_workbook(user["id"], user["report_master_name"])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/receipt-actions/close-month")
@login_required
def close_month():
    ensure_storage()
    user = current_user()
    try:
        output, filename = build_zip(user["id"], user["report_master_name"])
    except Exception as exc:
        return jsonify({"error": str(exc)}), 400
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/zip")


@app.post("/receipt-actions/reset")
@login_required
def reset_month():
    ensure_storage()
    user = current_user()
    user_receipts = db_rows(user["id"])
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute("DELETE FROM receipts WHERE user_id = ?", (user["id"],))
        conn.commit()
    for receipt in user_receipts:
        path = RECEIPT_DIR / receipt["stored_filename"]
        if path.exists():
            path.unlink()
    return jsonify({"ok": True})


if __name__ == "__main__":
    ensure_storage()
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)